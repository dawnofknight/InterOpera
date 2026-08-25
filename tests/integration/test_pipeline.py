from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

from interopera.orchestration.pipeline import run_pipeline

ROOT = Path(__file__).resolve().parents[2]
COMMON = {
    "guidelines": ROOT / "sample_docs/sample_fund_guidelines.pdf",
    "holdings": ROOT / "sample_docs/sample_holdings.csv",
    "template": ROOT / "sample_docs/report_template.xlsx",
    "approved_dir": ROOT / "approved_graph",
}


def test_firm_a_golden_and_replay(tmp_path) -> None:
    kwargs = dict(firm=ROOT / "config/firm_a.yaml", output=tmp_path,
        answer_key=ROOT / "sample_docs/firm_A_answer_key.xlsx", strict_reconcile=True, **COMMON)
    first = run_pipeline(**kwargs)
    second = run_pipeline(**kwargs)
    assert first["reconciliation"] == "PASS"
    assert first["figures_sha256"] == second["figures_sha256"]
    assert first["traces_sha256"] == second["traces_sha256"]
    figures = json.loads((Path(first["run_dir"]) / "computed_figures.json").read_text())
    assert len(figures) == 13


def test_firm_b_config_only_changes_expected_methods(tmp_path) -> None:
    result = run_pipeline(firm=ROOT / "config/firm_b.yaml", output=tmp_path,
                          strict_reconcile=False, **COMMON)
    figures = {row["figure_id"]: row for row in json.loads(
        (Path(result["run_dir"]) / "computed_figures.json").read_text())}
    assert figures["aggregate_non_ig_exposure"]["display_value"] == "21.0%"
    assert figures["aggregate_non_ig_exposure"]["status"] == "BREACH"
    assert figures["largest_gre_issuer"]["display_value"] == "13.0%"
    assert figures["largest_gre_issuer"]["status"] == "BREACH"
    assert figures["allocation_sgs"]["display_utilization"] == "5833 bps"
    assert figures["portfolio_dv01"]["display_value"] == "SGD 38,790 / bp"


def test_export_preserves_template_and_has_no_formulas(tmp_path) -> None:
    result = run_pipeline(firm=ROOT / "config/firm_a.yaml", output=tmp_path,
        answer_key=ROOT / "sample_docs/firm_A_answer_key.xlsx", strict_reconcile=True, **COMMON)
    report = Path(result["run_dir"]) / "firm_a_report.xlsx"
    source = load_workbook(ROOT / "sample_docs/report_template.xlsx")
    output = load_workbook(report, data_only=False)
    assert source.active.title == output.active.title
    for row in range(2, 15):
        assert source.active.cell(row, 2).value == output.active.cell(row, 2).value
        assert all(not (isinstance(output.active.cell(row, column).value, str) and
                        output.active.cell(row, column).value.startswith("=")) for column in range(3, 8))
    source.close()
    output.close()


def test_demo_resolves_assets_from_working_directory() -> None:
    """Installed modules must not look for config beside site-packages."""
    environment = dict(os.environ)
    environment["PYTHONPATH"] = str(ROOT / "src")
    completed = subprocess.run(
        [sys.executable, "-c", "import interopera.cli as c; print(c.ROOT)"],
        cwd=ROOT,
        env=environment,
        check=True,
        capture_output=True,
        text=True,
    )
    assert completed.stdout.strip() == str(ROOT)
