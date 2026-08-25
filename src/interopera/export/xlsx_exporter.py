from __future__ import annotations

import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from interopera.canonical import sha256_file
from interopera.domain.models import ComputedFigure


def export_xlsx(template: Path, destination: Path, figures: tuple[ComputedFigure, ...]) -> str:
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, destination)
    workbook = load_workbook(destination)
    sheet = workbook.active
    if sheet.max_row < 14 or sheet.max_column < 7:
        raise ValueError("Report template does not contain the required 13 rows and 7 columns")
    by_metric = {figure.metric: figure for figure in figures}
    for row in range(2, 15):
        metric = str(sheet.cell(row, 2).value)
        figure = by_metric.get(metric)
        if figure is None:
            raise ValueError(f"No computed figure for template metric: {metric}")
        values = (figure.display_value, figure.display_limit, figure.display_utilization,
                  figure.status.value, figure.source_summary)
        for column, value in enumerate(values, start=3):
            sheet.cell(row, column).value = value
    fixed = datetime(2000, 1, 1, tzinfo=timezone.utc)
    workbook.properties.created = fixed
    workbook.properties.modified = fixed
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.save(destination)
    workbook.close()
    return sha256_file(destination)

