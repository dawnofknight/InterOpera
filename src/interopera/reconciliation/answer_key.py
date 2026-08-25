from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def load_answer_key(path: Path) -> tuple[dict[str, str], ...]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    output = []
    for row in range(2, sheet.max_row + 1):
        output.append({
            "section": str(sheet.cell(row, 1).value), "metric": str(sheet.cell(row, 2).value),
            "value": str(sheet.cell(row, 3).value), "limit": str(sheet.cell(row, 4).value),
            "utilization": str(sheet.cell(row, 5).value), "status": str(sheet.cell(row, 6).value),
        })
    workbook.close()
    return tuple(output)

